"""Import vulnerabilities from CSV or Excel exports."""
from __future__ import annotations

import csv
import io
import re
from datetime import datetime, timezone
from typing import Any, BinaryIO, Optional

from packages.security.modules.vulnerabilities import db
from packages.security.modules.vulnerabilities.identity import (
    canonical_identity_key,
    import_external_id,
)
from packages.security.modules.vulnerabilities.remediation import prepare_import_fields

# Matches common scanner exports (e.g. Greenbone / OpenVAS style spreadsheets).
DEFAULT_COLUMN_MAP = {
    "vulnerability name": "title",
    "vulnerability": "title",
    "title": "title",
    "name": "title",
    "target": "ip",
    "ip": "ip",
    "target name": "host",
    "hostname": "host",
    "host": "host",
    "port": "port",
    "severity": "severity",
    "risk": "severity",
    "cvss": "cvss",
    "cve": "cve",
    "qod": "qod",
    "first occurrence": "first_seen",
    "first seen": "first_seen",
    "first_seen": "first_seen",
    "latest occurrence": "last_seen",
    "last seen": "last_seen",
    "last_seen": "last_seen",
    "solution": "solution",
    "solution external references": "refs_text",
    "proof": "description",
    "description": "description",
    "tags": "tags",
    "target criticality": "priority_hint",
    "comments": "import_comment",
}

SEVERITY_NORMALIZE = {
    "critical": "critical",
    "high": "high",
    "medium": "medium",
    "med": "medium",
    "low": "low",
    "info": "info",
    "informational": "info",
    "log": "info",
}


class ImportResult:
    def __init__(self) -> None:
        self.added = 0
        self.updated = 0
        self.merged = 0
        self.skipped = 0
        self.errors: list[str] = []

    @property
    def ok(self) -> bool:
        return not self.errors or (self.added + self.updated + self.merged) > 0

    def summary(self) -> str:
        parts = [
            f"{self.added} added",
            f"{self.updated} updated",
        ]
        if self.merged:
            parts.append(f"{self.merged} matched existing (no duplicate)")
        msg = "Import finished: " + ", ".join(parts)
        if self.skipped:
            msg += f", {self.skipped} skipped"
        if self.errors:
            msg += f", {len(self.errors)} issues"
        return msg


def _norm_header(h: str) -> str:
    return re.sub(r"\s+", " ", (h or "").strip().lower())


def _parse_float(val: Any) -> Optional[float]:
    if val is None or val == "":
        return None
    try:
        s = str(val).replace(",", ".").strip().rstrip("%")
        return float(s)
    except (TypeError, ValueError):
        return None


def _parse_severity(val: Any) -> str:
    s = str(val or "").strip().lower()
    if not s:
        return "info"
    if s in SEVERITY_NORMALIZE:
        return SEVERITY_NORMALIZE[s]
    score = _parse_float(s)
    if score is not None:
        if score >= 9.0:
            return "critical"
        if score >= 7.0:
            return "high"
        if score >= 4.0:
            return "medium"
        if score > 0:
            return "low"
    return "info"


def _cell_str(val: Any) -> str:
    if val is None:
        return ""
    if hasattr(val, "isoformat"):
        try:
            return val.isoformat()
        except Exception:
            pass
    return str(val).strip()


def _map_row(raw: dict[str, Any], header_map: dict[str, str]) -> dict[str, Any]:
    out: dict[str, Any] = {}
    for src_key, val in raw.items():
        field = header_map.get(_norm_header(src_key))
        if not field:
            continue
        out[field] = _cell_str(val) if val is not None and not isinstance(val, (int, float)) else val
    if not out.get("title"):
        out["title"] = out.get("cve") or out.get("host") or out.get("ip") or "Imported finding"
    sev = _parse_severity(out.get("severity"))
    out["severity"] = sev
    out["scanner_severity"] = sev
    cvss = _parse_float(out.get("cvss"))
    if cvss is not None:
        out["cvss"] = cvss
    if out.get("cve"):
        out["cve"] = str(out["cve"]).strip()
        out["kev"] = False
    tags_raw = out.pop("tags", "") or ""
    if tags_raw and str(tags_raw).lower() not in ("undefined", "[]", ""):
        parts = re.split(r"[,;]", str(tags_raw))
        out["tag_list"] = [t.strip() for t in parts if t.strip()]
    else:
        out["tag_list"] = []

    prepare_import_fields(out)
    identity_key = canonical_identity_key(
        cve=out.get("cve") or "",
        title=out.get("title") or "",
        host=out.get("host") or "",
        ip=out.get("ip") or "",
        port=out.get("port") or "",
    )
    out["identity_key"] = identity_key
    out["external_id"] = import_external_id(identity_key)
    out["raw"] = {**raw, "import_source": "csv_xlsx"}
    return out


def _upsert_mapped(row: dict, result: ImportResult) -> None:
    fields = {
        "external_id": row["external_id"],
        "identity_key": row["identity_key"],
        "title": str(row.get("title") or "")[:500],
        "severity": row.get("severity") or "info",
        "scanner_severity": row.get("scanner_severity") or row.get("severity") or "info",
        "cvss": row.get("cvss"),
        "host": str(row.get("host") or ""),
        "ip": str(row.get("ip") or ""),
        "port": str(row.get("port") or "") if row.get("port") not in (None, "") else "",
        "cve": str(row.get("cve") or ""),
        "description": str(row.get("description") or ""),
        "solution": str(row.get("solution") or ""),
        "refs": row.get("refs") or [],
        "first_seen": str(row.get("first_seen") or "") or datetime.now(timezone.utc).date().isoformat(),
        "last_seen": str(row.get("last_seen") or "") or datetime.now(timezone.utc).isoformat(),
        "raw": row.get("raw") or {},
    }
    vid, created, merged = db.upsert_vulnerability(fields)
    if row.get("tag_list"):
        db.set_tags(vid, row["tag_list"])
    comment = row.get("import_comment")
    if comment and str(comment).strip() and str(comment).lower() != "undefined":
        db.add_comment(vid, str(comment).strip(), 0, "import")
    if created:
        result.added += 1
    elif merged:
        result.merged += 1
    else:
        result.updated += 1


def _build_header_map(headers: list[str]) -> dict[str, str]:
    m: dict[str, str] = {}
    for h in headers:
        nh = _norm_header(h)
        if nh in DEFAULT_COLUMN_MAP:
            m[nh] = DEFAULT_COLUMN_MAP[nh]
    return m


def import_csv(file_obj: BinaryIO, *, encoding: str = "utf-8-sig") -> ImportResult:
    result = ImportResult()
    text = file_obj.read()
    if isinstance(text, bytes):
        text = text.decode(encoding, errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        result.errors.append("CSV has no header row.")
        return result
    header_map = _build_header_map(list(reader.fieldnames))
    if "title" not in header_map.values() and "cve" not in header_map.values():
        result.errors.append(
            "Unrecognized CSV columns. Expected headers like 'vulnerability name', 'target', 'severity', 'cve'."
        )
        return result
    for i, raw in enumerate(reader, start=2):
        try:
            mapped = _map_row(dict(raw), header_map)
            if not mapped.get("title"):
                result.skipped += 1
                continue
            _upsert_mapped(mapped, result)
        except Exception as e:
            result.errors.append(f"Row {i}: {e}")
    return result


def import_xlsx(file_obj: BinaryIO) -> ImportResult:
    result = ImportResult()
    try:
        from openpyxl import load_workbook
    except ImportError:
        result.errors.append("Excel support requires openpyxl (install requirements.txt).")
        return result
    wb = load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        headers = [str(c or "").strip() for c in next(rows)]
    except StopIteration:
        result.errors.append("Excel sheet is empty.")
        return result
    header_map = _build_header_map(headers)
    if "title" not in header_map.values() and "cve" not in header_map.values():
        result.errors.append(
            "Unrecognized Excel columns. Expected headers like 'vulnerability name', 'target', 'severity', 'cve'."
        )
        return result
    for i, row_vals in enumerate(rows, start=2):
        raw = {headers[j]: row_vals[j] if j < len(row_vals) else "" for j in range(len(headers))}
        if not any(v is not None and str(v).strip() for v in raw.values()):
            continue
        try:
            mapped = _map_row(raw, header_map)
            _upsert_mapped(mapped, result)
        except Exception as e:
            result.errors.append(f"Row {i}: {e}")
    wb.close()
    return result


def import_file(filename: str, file_obj: BinaryIO) -> ImportResult:
    lower = (filename or "").lower()
    if lower.endswith(".xlsx") or lower.endswith(".xlsm"):
        return import_xlsx(file_obj)
    return import_csv(file_obj)
